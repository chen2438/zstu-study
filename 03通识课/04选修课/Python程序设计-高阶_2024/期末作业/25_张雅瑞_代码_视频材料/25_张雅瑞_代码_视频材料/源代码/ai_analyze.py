import os
import openai
import pandas as pd
import json
from openpyxl import load_workbook
# optional; defaults to `os.environ['OPENAI_API_KEY']`
openai.api_key = "sk-BvqGnac4HYrcfMMYBd9c3268F3B142348e980589DaC28700"
# all client options can be configured just like the `OpenAI` instantiation counterpart
openai.base_url = "https://free.v36.cm/v1/"
openai.default_headers = {"x-foo": "true"}
def analyze_data(data_str):
    
    completion = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content":"你是数据分析专家和电影鉴赏专家，请你根据sheet的名称、用户提供的数据和互联网的评测，分析当前电源的形势，预测未来可能的电影趋势。并且你的回答要尽可能丰富且有理有据",
            },
            {
                "role": "user",
                "content":data_str,
            },
        ],
    )
    print(completion.choices[0].message.content)
    print("***********************************************\n")
    print("***********************************************\n")
    print("***********************************************\n")
    #print(completion.choices[0].message['content'])





def analyze_img(img):
    print("对于图表的分析")
    completion = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content":"你是数据分析专家和电影鉴赏专家，请你根据sheet的名称、用户提供的数据和互联网的评测，分析图表的信息并总结图表之间相互关联性。并且你的回答要尽可能丰富且有理有据",
            },
            {
                "role": "user",
                "content":img,
            },
        ],
    )
    print(completion.choices[0].message.content)
    print("***********************************************\n")
    print("***********************************************\n")
    print("***********************************************\n")
    #print(completion.choices[0].message['content'])





def sum_data(df):
    print("对数据的总结：")
    completion = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content":"你是数据分析专家和电影鉴赏专家，请你根据历史对话、excel提供的数据、用户提供的数据和互联网的评测，全部综合论述一下不同表格之间的关系，总结电影行业总的趋势",
            },
            {
                "role": "user",
                "content":df,
            },
        ],
    )
    print(completion.choices[0].message.content)     
    print("***********************************************\n")
    print("***********************************************\n")
    print("***********************************************\n")
        #print(completion.choices[0].message['content'])





def answer_data(df):
    while True:
        ask=input("请输入有关2020~2022年的电影相关的问题：")
        completion = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content":"你是数据分析专家和电影鉴赏专家，请你根据sheet的名称、用户提供的数据，回答用户的问题。注意，我们讨论的电影范围限定在2020~2022在表格中出现过的电影，你可以根据互联网数据调整你的回答，但是你的回答必须首先以用户提供的表格数据为准",
                },
                {
                    "role": "user",
                    "content":ask,
                },
                {
                    "role": "user",
                    "content":f"数据表格JSON: {df}",
                },
            ],
        )
        print(completion.choices[0].message.content)     
        print("***********************************************\n")
        print("***********************************************\n")
        print("***********************************************\n")
        #print(completion.choices[0].message['content'])





def excel_to_json(path):
    xls = pd.ExcelFile(path)
    data = {}
    for sheet_name in xls.sheet_names[:-1]:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        data[sheet_name] = df.to_dict(orient='records')
    json_data = json.dumps(data, ensure_ascii=False)
    return json_data





def extract_images(xls):
    # 使用 openpyxl 加载 Excel 文件
    workbook = load_workbook(xls.io, data_only=True)
    
    # 获取最后一个表格名称
    last_sheet_name = workbook.sheetnames[-1]
    last_sheet = workbook[last_sheet_name]
    images = []
    for image in last_sheet._images:
        images.append(image)  # 将图片对象添加到列表中
    print("true")
    return images




file_path = 'C:/Users/Mieopm/Desktop/top_movies_analysis_charts.xlsx'  # 替换为你的 Excel 文件路径
try:    
    xls = pd.ExcelFile(file_path)        
    for sheet_name in xls.sheet_names[:-1]:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        column_names = df.columns.tolist()
        sample_data = df.to_dict(orient='records')
        data = {
            "sheet_name":sheet_name,
            "columns": column_names,
            "data": sample_data
        }
        data_str = json.dumps(data, ensure_ascii=False)
        print("关于%s表格的数据分析如下："% (sheet_name))
        analyze_data(data_str)
    #imgs=extract_images(xls)
    #for img in imgs:
     #   analyze_img(img)
    sum_data(excel_to_json(file_path))
    answer_data(excel_to_json(file_path))
except FileNotFoundError:

    print(f"File not found: {file_path}")
except Exception as e:
    print(f"An error occurred: {e}")


